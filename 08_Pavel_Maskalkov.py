import csv
import openpyxl

with open('data.csv') as file:
    reader = csv.DictReader(file)
    lst_reader = list(reader)

    book = openpyxl.Workbook()
    sheet = book.active

    sheet['A2'] = 'id'
    sheet.cell(row=3, column=1).value = 'name'
    sheet[4][0].value = 'phone'

    column_labels = [f"person {i}" for i in range(1, len(lst_reader) + 1)]
    for col, label in zip(range(2, len(lst_reader) + 2), column_labels):
        sheet.cell(row=1, column=col).value = label

    for idx, row in enumerate(lst_reader, start=2):
        sheet.cell(row=2, column=idx).value = row['Id']
        sheet.cell(row=3, column=idx).value = row['Name']
        sheet.cell(row=4, column=idx).value = row['Phone']

    book.save('data.xlsx')
    book.close()